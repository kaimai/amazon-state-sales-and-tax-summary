"""
Generate Amazon order tax summary Excel from a raw Amazon order report TSV.

Usage:
    python amazon_tax_summary.py "/path/to/Amazon order report 2026 Jan.txt"

Output:
    "Amazon order tax summary - YYYYMM.xlsx" in the same directory as input.
    - Tab "YYYYMM detailed": all rows + "Sales Including Tax" column
    - Tab "YYYYMM summary": aggregated by country/state, all 51 US states shown
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

US_STATE_NAME_TO_ABBREV = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
    "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
    "DISTRICT OF COLUMBIA": "DC", "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI",
    "IDAHO": "ID", "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA",
    "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME",
    "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE",
    "NEVADA": "NV", "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM",
    "NEW YORK": "NY", "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH",
    "OKLAHOMA": "OK", "OREGON": "OR", "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI",
    "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX",
    "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA",
    "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

ALL_US_STATE_ABBREVS = sorted([
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DC", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA",
    "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY",
    "NC", "ND", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX",
    "UT", "VT", "VA", "WA", "WV", "WI", "WY",
])

BOLD = Font(bold=True)
CURRENCY_FMT = '#,##0.00'


def get_yyyymm(filepath: Path) -> str:
    name = filepath.stem
    m = re.search(r'(\d{4})\s+([A-Za-z]+)', name)
    if m:
        year, month_str = m.group(1), m.group(2)
        try:
            month_num = datetime.strptime(month_str[:3], "%b").month
            return f"{year}{month_num:02d}"
        except ValueError:
            pass
    # fallback: use earliest purchase-date in the file
    df = pd.read_csv(filepath, sep='\t', dtype=str, nrows=5)
    if 'purchase-date' in df.columns:
        dt = pd.to_datetime(df['purchase-date'].dropna().iloc[0], utc=True)
        return dt.strftime('%Y%m')
    raise ValueError(f"Cannot determine YYYYMM from filename: {name}")


def normalize_us_state(raw: str) -> str:
    if not raw or str(raw).strip() == '':
        return '(blank)'
    upper = str(raw).strip().upper()
    # already an abbreviation (2 chars)
    if len(upper) == 2:
        return upper
    return US_STATE_NAME_TO_ABBREV.get(upper, upper)


_NUMERIC_COLS = (
    'item-price', 'item-tax',
    'shipping-price', 'shipping-tax',
    'gift-wrap-price', 'gift-wrap-tax',
    'item-promotion-discount', 'ship-promotion-discount',
)


def load_data(filepath: Path) -> pd.DataFrame:
    df = pd.read_csv(filepath, sep='\t', dtype=str)
    # strip trailing whitespace from column names (Amazon sometimes adds spaces)
    df.columns = [c.strip() for c in df.columns]
    for col in _NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        else:
            df[col] = 0.0
    df['Sales Including Tax'] = (
        df['item-price'] + df['item-tax']
        + df['shipping-price'] + df['shipping-tax']
        + df['gift-wrap-price'] + df['gift-wrap-tax']
        - df['item-promotion-discount'] - df['ship-promotion-discount']
    )
    return df


def write_detailed_tab(ws, df: pd.DataFrame):
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    # format currency column
    sit_col = headers.index('Sales Including Tax') + 1
    for row in ws.iter_rows(min_row=2, min_col=sit_col, max_col=sit_col):
        for cell in row:
            cell.number_format = CURRENCY_FMT
    # auto-width for key columns
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(str(header)) + 2, 12)


def build_summary_rows(df: pd.DataFrame):
    """Return list of (country, state, sales_incl_tax, sales_tax) rows for the summary tab."""
    work = df.copy()
    work['ship-country'] = work['ship-country'].fillna('').str.strip()
    work['ship-state'] = work['ship-state'].fillna('').str.strip()

    # normalize US states
    us_mask = work['ship-country'].str.upper() == 'US'
    work.loc[us_mask, 'ship-state'] = work.loc[us_mask, 'ship-state'].apply(normalize_us_state)

    rows = []
    grand_sit = 0.0
    grand_tax = 0.0

    # --- International groups ---
    intl = work[~us_mask].copy()
    for country in sorted(intl['ship-country'].unique()):
        country_df = intl[intl['ship-country'] == country]
        country_sit = 0.0
        country_tax = 0.0
        for state in sorted(country_df['ship-state'].unique()):
            sdf = country_df[country_df['ship-state'] == state]
            sit = sdf['Sales Including Tax'].sum()
            tax = sdf['item-tax'].sum()
            state_label = state if state else '(blank)'
            rows.append((country, state_label, sit, tax))
            country_sit += sit
            country_tax += tax
        rows.append((f"{country} Total", '', country_sit, country_tax))
        grand_sit += country_sit
        grand_tax += country_tax

    # --- US section ---
    us = work[us_mask].copy()
    us_sit = us.groupby('ship-state')['Sales Including Tax'].sum()
    us_tax = us.groupby('ship-state')['item-tax'].sum()
    us_total_sit = 0.0
    us_total_tax = 0.0
    for state in ALL_US_STATE_ABBREVS:
        sit = float(us_sit.get(state, 0.0))
        tax = float(us_tax.get(state, 0.0))
        rows.append(('US', state, sit, tax))
        us_total_sit += sit
        us_total_tax += tax
    rows.append(('US Total', '', us_total_sit, us_total_tax))
    grand_sit += us_total_sit
    grand_tax += us_total_tax

    rows.append(('Grand Total', '', grand_sit, grand_tax))
    return rows


def write_summary_tab(ws, summary_rows):
    header = ['ship-country', 'ship-state', 'SUM of Sales Including Tax', 'SUM of Sales Tax']
    ws.append(header)
    for cell in ws[1]:
        cell.font = BOLD

    for country, state, sit, tax in summary_rows:
        ws.append([country, state, sit, tax])
        row_idx = ws.max_row
        is_total = str(country).endswith(' Total') or country == 'Grand Total'
        if is_total:
            for cell in ws[row_idx]:
                cell.font = BOLD
        ws.cell(row=row_idx, column=3).number_format = CURRENCY_FMT
        ws.cell(row=row_idx, column=4).number_format = CURRENCY_FMT

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20


def main(input_path: str, append_to: str | None = None):
    from openpyxl import load_workbook

    filepath = Path(input_path)
    if not filepath.exists():
        print(f"ERROR: file not found: {filepath}")
        sys.exit(1)

    yyyymm = get_yyyymm(filepath)
    print(f"Detected period: {yyyymm}")

    df = load_data(filepath)
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

    if append_to:
        out_path = Path(append_to)
        wb = load_workbook(out_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        out_path = filepath.parent / f"Amazon state sales and tax summary - {yyyymm}.xlsx"

    # Remove existing tabs for this period if re-running
    for name in (f"{yyyymm} detailed", f"{yyyymm} summary"):
        if name in wb.sheetnames:
            del wb[name]

    # Tab 1: detailed
    detailed_tab_name = f"{yyyymm} detailed"
    ws_detail = wb.create_sheet(title=detailed_tab_name)
    write_detailed_tab(ws_detail, df)
    print(f"Wrote '{detailed_tab_name}' tab: {len(df)} data rows")

    # Tab 2: summary
    summary_tab_name = f"{yyyymm} summary"
    ws_summary = wb.create_sheet(title=summary_tab_name)
    summary_rows = build_summary_rows(df)
    write_summary_tab(ws_summary, summary_rows)
    us_state_rows = sum(1 for c, s, sit, tax in summary_rows if c == 'US')
    print(f"Wrote '{summary_tab_name}' tab: {us_state_rows} US states")

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('input', help='Path to Amazon order report TSV')
    parser.add_argument('--append-to', help='Append tabs to this existing Excel file')
    args = parser.parse_args()
    main(args.input, append_to=args.append_to)
