"""
Integration tests for amazon_tax_summary.py.

Runs the script against fixture input files and compares output against
expected CSV snapshots.
"""

from __future__ import annotations

import csv
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
import pandas as pd

# Make the root importable regardless of where pytest is invoked from
sys.path.insert(0, str(Path(__file__).parent.parent))
import amazon_tax_summary as ats

FIXTURES = Path(__file__).parent / "fixtures"
INPUT_DIR = FIXTURES / "input"
EXPECTED_DIR = FIXTURES / "expected"


def read_summary_tab(xlsx_path: Path, tab_name: str) -> list[tuple]:
    """Return all non-empty rows from a summary tab as (country, state, sit, tax) tuples."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[tab_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        country, state, sit, tax = row[0], row[1], row[2], row[3]
        if country is not None:
            rows.append((country, state or "", round(float(sit), 2), round(float(tax), 2)))
    return rows


def load_expected_csv(path: Path) -> list[tuple]:
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sit = round(float(row["SUM of Sales Including Tax"]), 2)
            tax = round(float(row["SUM of Sales Tax"]), 2)
            rows.append((row["ship-country"], row["ship-state"], sit, tax))
    return rows


class TestJanuaryReport(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        cls.out = Path(cls.tmp) / "Amazon order tax summary - 202601.xlsx"
        # Copy input to tmp so output lands next to it
        import shutil
        src = INPUT_DIR / "Amazon order report 2026 Jan.txt"
        dst = Path(cls.tmp) / src.name
        shutil.copy(src, dst)
        ats.main(str(dst))

    # ── Detailed tab ──────────────────────────────────────────────────────────

    def test_detailed_tab_exists(self):
        wb = openpyxl.load_workbook(self.out)
        self.assertIn("202601 detailed", wb.sheetnames)

    def test_detailed_row_count(self):
        wb = openpyxl.load_workbook(self.out)
        ws = wb["202601 detailed"]
        # 1 header + 10 data rows (including cancelled)
        self.assertEqual(ws.max_row, 11)

    def test_detailed_has_sales_including_tax_column(self):
        wb = openpyxl.load_workbook(self.out)
        ws = wb["202601 detailed"]
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Sales Including Tax", headers)

    def test_sales_including_tax_computed_correctly(self):
        # Row 2: item-price=100, item-tax=8.50 → 108.50
        wb = openpyxl.load_workbook(self.out)
        ws = wb["202601 detailed"]
        headers = [cell.value for cell in ws[1]]
        sit_col = headers.index("Sales Including Tax") + 1
        first_data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertAlmostEqual(first_data_row[sit_col - 1], 108.50, places=2)

    def test_cancelled_order_included_in_detailed(self):
        # Cancelled row has 0.00 Sales Including Tax but must be present
        wb = openpyxl.load_workbook(self.out)
        ws = wb["202601 detailed"]
        statuses = [row[4] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("Cancelled", statuses)

    # ── Summary tab ───────────────────────────────────────────────────────────

    def test_summary_tab_exists(self):
        wb = openpyxl.load_workbook(self.out)
        self.assertIn("202601 summary", wb.sheetnames)

    def test_summary_matches_expected(self):
        actual = read_summary_tab(self.out, "202601 summary")
        expected = load_expected_csv(EXPECTED_DIR / "202601_summary.csv")
        self.assertEqual(actual, expected)

    def test_all_52_us_entries_present(self):
        actual = read_summary_tab(self.out, "202601 summary")
        us_states = [s for c, s, sit, tax in actual if c == "US"]
        self.assertEqual(len(us_states), 52)

    def test_california_aggregated_from_abbrev_and_full_name(self):
        # Rows 1+2 use "CA", row 10 uses "CALIFORNIA" — all should sum to 475.50
        actual = read_summary_tab(self.out, "202601 summary")
        ca_sit = next(sit for c, s, sit, tax in actual if c == "US" and s == "CA")
        self.assertAlmostEqual(ca_sit, 475.50, places=2)
        ca_tax = next(tax for c, s, sit, tax in actual if c == "US" and s == "CA")
        self.assertAlmostEqual(ca_tax, 25.50, places=2)

    def test_maryland_normalized_from_full_name(self):
        actual = read_summary_tab(self.out, "202601 summary")
        md_sit = next(sit for c, s, sit, tax in actual if c == "US" and s == "MD")
        self.assertAlmostEqual(md_sit, 418.15, places=2)
        md_tax = next(tax for c, s, sit, tax in actual if c == "US" and s == "MD")
        self.assertAlmostEqual(md_tax, 19.15, places=2)

    def test_puerto_rico_in_us_section(self):
        actual = read_summary_tab(self.out, "202601 summary")
        pr_sit = next(sit for c, s, sit, tax in actual if c == "US" and s == "PR")
        self.assertAlmostEqual(pr_sit, 100.00, places=2)

    def test_cancelled_order_contributes_zero_to_summary(self):
        # WA had only a cancelled order → should be 0.00
        actual = read_summary_tab(self.out, "202601 summary")
        wa_sit = next(sit for c, s, sit, tax in actual if c == "US" and s == "WA")
        self.assertAlmostEqual(wa_sit, 0.00, places=2)

    def test_international_canada_grouped(self):
        actual = read_summary_tab(self.out, "202601 summary")
        ca_total_sit = next(sit for c, s, sit, tax in actual if c == "CA Total")
        self.assertAlmostEqual(ca_total_sit, 299.00, places=2)
        ca_total_tax = next(tax for c, s, sit, tax in actual if c == "CA Total")
        self.assertAlmostEqual(ca_total_tax, 0.00, places=2)

    def test_us_total(self):
        actual = read_summary_tab(self.out, "202601 summary")
        us_total_sit = next(sit for c, s, sit, tax in actual if c == "US Total")
        self.assertAlmostEqual(us_total_sit, 1515.12, places=2)
        us_total_tax = next(tax for c, s, sit, tax in actual if c == "US Total")
        self.assertAlmostEqual(us_total_tax, 87.12, places=2)

    def test_grand_total(self):
        actual = read_summary_tab(self.out, "202601 summary")
        grand_sit = next(sit for c, s, sit, tax in actual if c == "Grand Total")
        self.assertAlmostEqual(grand_sit, 1814.12, places=2)
        grand_tax = next(tax for c, s, sit, tax in actual if c == "Grand Total")
        self.assertAlmostEqual(grand_tax, 87.12, places=2)


class TestAppendMode(unittest.TestCase):
    def test_append_adds_new_month_tabs(self):
        tmp = tempfile.mkdtemp()
        import shutil
        src = INPUT_DIR / "Amazon order report 2026 Jan.txt"
        dst = Path(tmp) / src.name
        shutil.copy(src, dst)

        # First: create Jan file
        ats.main(str(dst))
        jan_out = Path(tmp) / "Amazon order tax summary - 202601.xlsx"

        # Simulate a Feb input with a single row
        feb_tsv = Path(tmp) / "Amazon order report 2026 Feb.txt"
        feb_tsv.write_text(
            "amazon-order-id\titem-price\titem-tax\tship-state\tship-country\n"
            "222-0000001\t50.00\t4.00\tNY\tUS\n"
        )
        ats.main(str(feb_tsv), append_to=str(jan_out))

        wb = openpyxl.load_workbook(jan_out)
        self.assertIn("202601 detailed", wb.sheetnames)
        self.assertIn("202601 summary", wb.sheetnames)
        self.assertIn("202602 detailed", wb.sheetnames)
        self.assertIn("202602 summary", wb.sheetnames)

    def test_rerun_overwrites_existing_tabs(self):
        tmp = tempfile.mkdtemp()
        import shutil
        src = INPUT_DIR / "Amazon order report 2026 Jan.txt"
        dst = Path(tmp) / src.name
        shutil.copy(src, dst)
        ats.main(str(dst))
        out = Path(tmp) / "Amazon order tax summary - 202601.xlsx"

        ats.main(str(dst), append_to=str(out))

        wb = openpyxl.load_workbook(out)
        self.assertEqual(wb.sheetnames.count("202601 detailed"), 1)
        self.assertEqual(wb.sheetnames.count("202601 summary"), 1)


if __name__ == "__main__":
    unittest.main()
